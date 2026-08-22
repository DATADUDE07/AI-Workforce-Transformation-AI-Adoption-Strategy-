"""
08_excel_export.py
------------------
Generates the 8-sheet business-facing Excel workbook:
  AI_Workforce_Strategy.xlsx

Uses openpyxl (standard Excel writing library) for:
  - Multi-sheet workbook
  - RAG conditional formatting (green/amber/red cell fills)
  - Bold headers, column widths, frozen panes
  - Embedded chart images
  - Executive Summary narrative block

openpyxl is the standard Python library for .xlsx output.
No xlwt or xlrd needed (those are legacy .xls formats).
"""

import os
import sys
import numpy as np
import pandas as pd
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    numbers as openpyxl_numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config
import utils

utils.ensure_dirs()

# --- Style constants ----------------------------------------------------------
HEADER_FONT    = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT     = Font(name="Calibri", bold=True, size=13, color="1A3C5E")
LABEL_FONT     = Font(name="Calibri", bold=True, size=10, color="1A3C5E")
BODY_FONT      = Font(name="Calibri", size=10)
SMALL_FONT     = Font(name="Calibri", size=9, italic=True)

HEADER_FILL    = PatternFill("solid", fgColor="1A3C5E")
GREEN_FILL     = PatternFill("solid", fgColor="27AE60")
AMBER_FILL     = PatternFill("solid", fgColor="F39C12")
RED_FILL       = PatternFill("solid", fgColor="C0392B")
LIGHT_BLUE     = PatternFill("solid", fgColor="D6EAF8")
LIGHT_GREY     = PatternFill("solid", fgColor="F2F3F4")
WHITE_FILL     = PatternFill("solid", fgColor="FFFFFF")

CENTER_ALIGN   = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT_ALIGN    = Alignment(horizontal="right",  vertical="center")

THIN_BORDER    = Border(
    left=Side(style="thin",   color="BDC3C7"),
    right=Side(style="thin",  color="BDC3C7"),
    top=Side(style="thin",    color="BDC3C7"),
    bottom=Side(style="thin", color="BDC3C7"),
)

NUMBER_FMT_USD   = '#,##0'
NUMBER_FMT_MUSD  = '#,##0.00'
NUMBER_FMT_PCT   = '0.0%'
NUMBER_FMT_SCORE = '0.000'


# ══════════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

def write_header_row(ws, row_num: int, headers: list, start_col: int = 1):
    for j, h in enumerate(headers, start=start_col):
        cell = ws.cell(row=row_num, column=j, value=h)
        cell.font   = HEADER_FONT
        cell.fill   = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def write_data_row(ws, row_num: int, values: list, start_col: int = 1, alt: bool = False):
    fill = LIGHT_GREY if alt else WHITE_FILL
    for j, v in enumerate(values, start=start_col):
        cell = ws.cell(row=row_num, column=j, value=v)
        cell.font      = BODY_FONT
        cell.fill      = fill
        cell.alignment = RIGHT_ALIGN if isinstance(v, (int, float)) else LEFT_ALIGN
        cell.border    = THIN_BORDER


def rag_fill(value: str) -> PatternFill:
    fills = {"Green": GREEN_FILL, "Amber": AMBER_FILL, "Red": RED_FILL}
    return fills.get(value, WHITE_FILL)


def set_col_widths(ws, widths: list, start_col: int = 1):
    for j, w in enumerate(widths, start=start_col):
        ws.column_dimensions[get_column_letter(j)].width = w


def fmt_musd(v) -> str:
    try:
        return f"${float(v)/1e6:.2f}M"
    except Exception:
        return str(v)


def fmt_pct(v) -> str:
    try:
        return f"{float(v)*100:.1f}%"
    except Exception:
        return str(v)


def fmt_usd_k(v) -> str:
    try:
        return f"${float(v)/1e3:.0f}K"
    except Exception:
        return str(v)


def title_block(ws, row: int, col: int, text: str, note: str = ""):
    ws.cell(row=row, column=col, value=text).font = TITLE_FONT
    if note:
        c = ws.cell(row=row+1, column=col, value=note)
        c.font = SMALL_FONT


def insert_image(ws, img_path: str, anchor: str, width_px: int = 520, height_px: int = 290):
    """Embed a chart PNG into the worksheet at the given anchor cell."""
    if os.path.exists(img_path):
        img = XLImage(img_path)
        img.width  = width_px
        img.height = height_px
        ws.add_image(img, anchor)
    else:
        ws[anchor] = f"[Chart not found: {os.path.basename(img_path)}]"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — EXECUTIVE SUMMARY
# ══════════════════════════════════════════════════════════════════════════════

def sheet_executive_summary(wb, master, scoring, scenarios, summary):
    ws = wb.create_sheet("1. Executive Summary")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B3"

    # Company header
    ws.merge_cells("B2:L2")
    c = ws["B2"]
    c.value = "AI WORKFORCE TRANSFORMATION STRATEGY — EXECUTIVE SUMMARY"
    c.font  = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    c.fill  = HEADER_FILL
    c.alignment = CENTER_ALIGN
    ws.row_dimensions[2].height = 35

    # Timestamp
    ws["B3"] = f"Analysis Date: {datetime.now().strftime('%B %Y')}  |  " \
               f"Synthetic case study — all data illustrative"
    ws["B3"].font = SMALL_FONT
    ws["B3"].alignment = LEFT_ALIGN

    # -- KPI boxes ----------------------------------------------------------
    mod_summary = summary[summary["scenario"] == "Moderate"].iloc[0]
    agg_summary = summary[summary["scenario"] == "Aggressive"].iloc[0]
    con_summary = summary[summary["scenario"] == "Conservative"].iloc[0]

    total_labor   = master["total_annual_labor_cost_usd"].sum()
    total_heads   = master["total_headcount"].sum()
    green_roles   = (scoring["RAG"] == "Green").sum()
    phase1_count  = (scoring["Phase"] == "Phase 1").sum()
    mod_savings   = mod_summary["total_labor_savings"]
    mod_roi       = mod_summary["portfolio_roi"]
    mod_payback   = mod_summary["portfolio_payback_months"]
    mod_npv       = mod_summary["total_npv_3yr"]

    kpis = [
        ("Total Workforce Analysed",  f"{int(total_heads):,} FTE",              "B5"),
        ("Total Annual Labor Cost",   f"${total_labor/1e6:.1f}M",               "E5"),
        ("Moderate Scenario Savings", f"${mod_savings/1e6:.1f}M/yr",            "H5"),
        ("Portfolio ROI (Moderate)",  f"{mod_roi*100:.0f}%",                    "K5"),
        ("Payback Period",            f"{mod_payback:.0f} months",              "B9"),
        ("3-Year NPV (Moderate)",     f"${mod_npv/1e6:.1f}M",                   "E9"),
        ("Green Priority Roles",      f"{green_roles} of 30 roles",             "H9"),
        ("Phase 1 Roles",             f"{phase1_count} roles ready now",        "K9"),
    ]
    for label, value, anchor in kpis:
        col = openpyxl.utils.cell.coordinate_to_tuple(anchor)[1]
        row = openpyxl.utils.cell.coordinate_to_tuple(anchor)[0]
        ws.merge_cells(start_row=row, start_column=col, end_row=row+1, end_column=col+2)
        ws.merge_cells(start_row=row+2, start_column=col, end_row=row+2, end_column=col+2)
        c1 = ws.cell(row=row, column=col, value=label)
        c1.font = Font(name="Calibri", bold=True, size=10, color="1A3C5E")
        c1.fill = LIGHT_BLUE
        c1.alignment = CENTER_ALIGN
        c1.border = THIN_BORDER
        c2 = ws.cell(row=row+2, column=col, value=value)
        c2.font = Font(name="Calibri", bold=True, size=14, color="1A3C5E")
        c2.fill = WHITE_FILL
        c2.alignment = CENTER_ALIGN
        c2.border = THIN_BORDER

    # -- Key Findings narrative ---------------------------------------------
    ws.merge_cells("B14:L14")
    h = ws["B14"]
    h.value = "KEY FINDINGS"
    h.font  = LABEL_FONT
    h.fill  = LIGHT_BLUE
    h.alignment = LEFT_ALIGN

    findings = [
        "1. Customer Support (Tier 1) and Finance Operations (AP/AR, Payroll) score highest on the AI Adoption Priority Score (AAPS) — high repetitiveness, strong AI tool maturity, and large cost pools.",
        f"2. Under the Moderate AI scenario, estimated annual labor savings of ${mod_savings/1e6:.1f}M are achievable against a total labor cost base of ${total_labor/1e6:.1f}M ({mod_savings/total_labor:.1%} reduction).",
        f"3. The portfolio investment of ${mod_summary['total_impl_cost']/1e6:.1f}M generates a Year-1 ROI of {mod_roi*100:.0f}% and a payback period of {mod_payback:.0f} months under the moderate scenario.",
        f"4. A 3-year NPV of ${mod_npv/1e6:.1f}M (at 10% discount rate) validates the business case across all three scenarios.",
        "5. Roles such as Corporate Counsel, HR Business Partner, and Account Executive (Enterprise) should NOT be automated — their decision intensity and client relationship value far exceeds any automation benefit.",
        "6. Sensitivity analysis confirms that AAPS rankings are robust (Spearman rho >= 0.85) — the top-priority roles remain consistent regardless of weight assumptions.",
        "7. RECOMMENDED STRATEGY: Moderate scenario, phased 24-month rollout, starting with Finance Ops and Customer Support before expanding to Engineering (DevOps/QA) and Data & Analytics.",
    ]
    for i, txt in enumerate(findings):
        r = 15 + i
        ws.merge_cells(f"B{r}:L{r}")
        c = ws[f"B{r}"]
        c.value = txt
        c.font = BODY_FONT
        c.alignment = LEFT_ALIGN
        c.border = THIN_BORDER
        c.fill = LIGHT_GREY if i % 2 == 0 else WHITE_FILL

    # -- Scenario comparison table ------------------------------------------
    title_block(ws, 24, 2, "Scenario Comparison at a Glance")
    headers = ["Scenario", "Labor Savings", "Impl. Cost", "Net Annual Benefit",
               "Savings %", "Portfolio ROI", "Payback (mo)", "3-yr NPV", "Risk Level"]
    write_header_row(ws, 25, headers, start_col=2)
    sc_risks = {"Conservative": "Low", "Moderate": "Medium", "Aggressive": "High"}

    for i, (_, row) in enumerate(summary.iterrows()):
        vals = [
            row["scenario"],
            fmt_musd(row["total_labor_savings"]),
            fmt_musd(row["total_impl_cost"]),
            fmt_musd(row["total_net_annual_benefit"]),
            fmt_pct(row["savings_pct_of_labor"]),
            f"{row['portfolio_roi']*100:.0f}%",
            f"{row['portfolio_payback_months']:.0f}",
            fmt_musd(row["total_npv_3yr"]),
            sc_risks[row["scenario"]],
        ]
        write_data_row(ws, 26 + i, vals, start_col=2, alt=i % 2 == 0)

    # Embed chart: AAPS ranking
    insert_image(ws, os.path.join(config.CHARTS_DIR, "09_risk_vs_opportunity.png"), "B31", 560, 300)

    set_col_widths(ws, [2, 22, 16, 16, 20, 14, 16, 16, 14, 16, 14], start_col=1)
    ws.row_dimensions[2].height = 36
    for r in range(5, 12):
        ws.row_dimensions[r].height = 22


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — WORKFORCE & COST ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════

def sheet_workforce_cost(wb, master):
    ws = wb.create_sheet("2. Workforce & Cost")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    title_block(ws, 1, 1, "Workforce & Labor Cost Analysis",
                "All costs include 30% benefits load. Salaries anchored to BLS/Glassdoor public benchmarks (2023-24).")

    headers = ["Role", "Department", "Level", "Headcount", "Avg Salary ($)",
               "Benefits Load", "Total Labor Cost ($)", "Cost / Hour ($)",
               "Productive Hrs/yr", "Cost % of Total"]
    write_header_row(ws, 3, headers)

    total_cost = master["total_annual_labor_cost_usd"].sum()
    for i, (_, row) in enumerate(master.sort_values("total_annual_labor_cost_usd", ascending=False).iterrows()):
        vals = [
            row["role_name"], row["function"], row["role_level"],
            int(row["total_headcount"]), int(row["avg_annual_salary_usd"]),
            f"30%",
            int(row["total_annual_labor_cost_usd"]),
            round(row.get("cost_per_productive_hour", 0), 2),
            int(row.get("productive_hours_per_year", config.HOURS_PER_WEEK * config.WORKING_WEEKS)),
            f"{row['total_annual_labor_cost_usd']/total_cost*100:.1f}%",
        ]
        write_data_row(ws, 4 + i, vals, alt=i % 2 == 0)

    # Department subtotals
    start = 4 + len(master) + 2
    ws.cell(row=start, column=1, value="DEPARTMENT SUBTOTALS").font = LABEL_FONT
    by_dept = master.groupby("function").agg(
        Headcount=("total_headcount", "sum"),
        Total_Cost=("total_annual_labor_cost_usd", "sum"),
    ).sort_values("Total_Cost", ascending=False)

    write_header_row(ws, start + 1, ["Department", "Headcount", "Total Cost ($)", "Cost Share (%)"])
    for i, (dept, row) in enumerate(by_dept.iterrows()):
        vals = [dept, int(row["Headcount"]), int(row["Total_Cost"]),
                f"{row['Total_Cost']/total_cost*100:.1f}%"]
        write_data_row(ws, start + 2 + i, vals, alt=i % 2 == 0)

    # Embed chart
    insert_image(ws, os.path.join(config.CHARTS_DIR, "01_labor_cost_by_department.png"),
                 f"L2", 520, 280)

    set_col_widths(ws, [30, 18, 10, 12, 16, 14, 20, 16, 18, 14], start_col=1)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — AI AUTOMATION ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════

def sheet_ai_automation(wb, master):
    ws = wb.create_sheet("3. AI Automation Analysis")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    title_block(ws, 1, 1, "AI Automation & Augmentation Analysis",
                "Automation/augmentation potentials derived from role-logic rules. See data provenance notes.")

    headers = [
        "Role", "Department", "AI Classification",
        "Auto Potential %", "Aug Potential %", "Human Oversight %",
        "AI Readiness %", "Automatable Cost Pool ($)",
        "Productivity Uplift %", "Time to Value (mo)",
    ]
    write_header_row(ws, 3, headers)

    class_fill_map = {
        "Full Automation":    PatternFill("solid", fgColor="D5F5E3"),
        "AI Augmentation":   PatternFill("solid", fgColor="FEF9E7"),
        "Low AI Suitability": PatternFill("solid", fgColor="FADBD8"),
    }

    for i, (_, row) in enumerate(master.sort_values("automation_potential_pct", ascending=False).iterrows()):
        vals = [
            row["role_name"], row["function"],
            row.get("ai_classification", ""),
            f"{row['automation_potential_pct']*100:.1f}%",
            f"{row['augmentation_potential_pct']*100:.1f}%",
            f"{row.get('human_oversight_required_pct', 0)*100:.1f}%",
            f"{row['ai_maturity_readiness']*100:.1f}%",
            int(row.get("automatable_labor_cost_usd", 0)),
            f"{row['ai_productivity_uplift_pct']*100:.1f}%",
            int(row.get("time_to_value_months", 0)),
        ]
        write_data_row(ws, 4 + i, vals, alt=i % 2 == 0)
        cls = row.get("ai_classification", "")
        if cls in class_fill_map:
            ws.cell(row=4 + i, column=3).fill = class_fill_map[cls]

    insert_image(ws, os.path.join(config.CHARTS_DIR, "02_automation_potential_by_function.png"),
                 "L2", 540, 290)
    insert_image(ws, os.path.join(config.CHARTS_DIR, "06_auto_vs_judgment.png"),
                 "L22", 540, 290)

    set_col_widths(ws, [30, 18, 18, 16, 16, 18, 16, 24, 20, 18])


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — ROLE / TASK SCORING
# ══════════════════════════════════════════════════════════════════════════════

def sheet_role_scoring(wb, scoring):
    ws = wb.create_sheet("4. Role & Task Scoring")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    title_block(ws, 1, 1, "AI Adoption Priority Score (AAPS) — Full Scoring Table",
                f"AAPS = 0.25xAuto + 0.25xCostSaving + 0.15xProductivity + 0.15xFeasibility + 0.10x(1−Risk) + 0.10x(1−Judgment)")

    headers = [
        "Rank", "Role", "Department", "Phase", "RAG",
        "AAPS Score", "F1: Auto Pot", "F2: Cost Saving", "F3: Productivity",
        "F4: Feasibility", "Risk Score", "AI Classification",
    ]
    write_header_row(ws, 3, headers)

    df_sorted = scoring.sort_values("AAPS_rank")
    for i, (_, row) in enumerate(df_sorted.iterrows()):
        vals = [
            int(row["AAPS_rank"]),
            row["role_name"],
            row["function"],
            str(row.get("Phase", "")),
            row.get("RAG", ""),
            round(float(row["AAPS"]), 4),
            round(float(row["F1_norm"]), 3),
            round(float(row["F2_norm"]), 3),
            round(float(row["F3_norm"]), 3),
            round(float(row["F4_norm"]), 3),
            round(float(row.get("risk_score", 0)), 3),
            str(row.get("ai_classification", "")),
        ]
        write_data_row(ws, 4 + i, vals, alt=i % 2 == 0)

        # RAG fill on RAG cell
        rag_val = row.get("RAG", "")
        ws.cell(row=4+i, column=5).fill = rag_fill(rag_val)
        ws.cell(row=4+i, column=5).font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

    insert_image(ws, os.path.join(config.CHARTS_DIR, "05_aaps_ranking.png"), "N2", 560, 580)

    set_col_widths(ws, [6, 30, 18, 10, 8, 10, 12, 12, 14, 14, 10, 20])


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — SCENARIO ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════

def sheet_scenario_analysis(wb, scenarios, summary, master):
    ws = wb.create_sheet("5. Scenario Analysis")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    title_block(ws, 1, 1, "Scenario Analysis — Conservative / Moderate / Aggressive",
                "Adoption rates: Conservative 40%, Moderate 65%, Aggressive 85% of automatable tasks captured.")

    # Scenario parameter table
    param_headers = ["Parameter", "Conservative", "Moderate", "Aggressive"]
    write_header_row(ws, 3, param_headers)
    params_data = [
        ("Adoption Rate",               "40%", "65%", "85%"),
        ("Productivity Mult.",          "50%", "75%", "90%"),
        ("Impl. Cost Multiplier",       "1.20x", "1.00x", "0.90x"),
        ("Ongoing Cost Multiplier",     "1.00x", "1.00x", "1.10x"),
        ("Risk Materialisation",        "80%", "60%", "40%"),
        ("Rollout Timeline",            "36 mo", "24 mo", "18 mo"),
    ]
    for i, vals in enumerate(params_data):
        write_data_row(ws, 4 + i, list(vals), alt=i % 2 == 0)

    # Summary comparison
    title_block(ws, 12, 1, "Company-Level Financial Summary")
    sum_headers = ["Metric", "Conservative", "Moderate", "Aggressive"]
    write_header_row(ws, 13, sum_headers)
    rows_map = [
        ("Labor Savings", "total_labor_savings"),
        ("Productivity Value", "total_productivity_value"),
        ("Implementation Cost", "total_impl_cost"),
        ("Ongoing AI Cost", "total_ongoing_cost"),
        ("Net Annual Benefit", "total_net_annual_benefit"),
        ("Savings % of Labor", "savings_pct_of_labor"),
        ("Portfolio ROI (Yr1)", "portfolio_roi"),
        ("Payback Period", "portfolio_payback_months"),
        ("3-Year NPV", "total_npv_3yr"),
        ("Risk-Adj. Benefit", "total_risk_adj_benefit"),
    ]
    sc_data = {row["scenario"]: row for _, row in summary.iterrows()}
    for i, (label, key) in enumerate(rows_map):
        c_val = sc_data.get("Conservative", {}).get(key, 0)
        m_val = sc_data.get("Moderate",     {}).get(key, 0)
        a_val = sc_data.get("Aggressive",   {}).get(key, 0)

        def fmt(v, key):
            if "pct" in key or "roi" in key:
                return f"{float(v)*100:.1f}%"
            elif "months" in key:
                return f"{float(v):.0f} mo"
            else:
                return fmt_musd(v)

        vals = [label, fmt(c_val, key), fmt(m_val, key), fmt(a_val, key)]
        write_data_row(ws, 14 + i, vals, alt=i % 2 == 0)

    # Role-level moderate scenario
    title_block(ws, 27, 1, "Per-Role Results — Moderate Scenario")
    role_headers = ["Role", "Department", "Labor Savings ($)", "Prod. Value ($)",
                    "Net Benefit ($)", "ROI Year-1", "Payback (mo)", "3-yr NPV ($)"]
    write_header_row(ws, 28, role_headers)
    mod = scenarios[scenarios["scenario"] == "Moderate"].sort_values("net_annual_benefit", ascending=False)
    for i, (_, row) in enumerate(mod.iterrows()):
        vals = [
            row["role_name"], row["function"],
            int(row["labor_savings"]), int(row["productivity_gain_value"]),
            int(row["net_annual_benefit"]),
            f"{row['roi_year1']*100:.0f}%",
            f"{row['payback_months']:.0f}" if row["payback_months"] < 1000 else "N/A",
            int(row["npv_3yr"]),
        ]
        write_data_row(ws, 29 + i, vals, alt=i % 2 == 0)

    insert_image(ws, os.path.join(config.CHARTS_DIR, "07_scenario_savings_by_department.png"),
                 "F2", 520, 300)

    set_col_widths(ws, [32, 18, 18, 14, 16, 14, 14, 16])


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — ROI & SAVINGS
# ══════════════════════════════════════════════════════════════════════════════

def sheet_roi_savings(wb, scenarios, master):
    ws = wb.create_sheet("6. ROI & Savings")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    title_block(ws, 1, 1, "ROI & Savings Analysis",
                "Formulas: ROI = (Net Benefit − Impl. Cost) / Impl. Cost  |  NPV at 10% discount rate  |  Moderate scenario")

    headers = [
        "Rank", "Role", "Department",
        "Impl. Cost ($K)", "Net Benefit/yr ($K)",
        "ROI Year-1 (%)", "Payback (mo)", "3-yr NPV ($K)",
        "Labor Savings ($K)", "Savings % of Cost",
    ]
    write_header_row(ws, 3, headers)

    mod = scenarios[scenarios["scenario"] == "Moderate"].copy()
    # total_annual_labor_cost_usd is already in the scenarios CSV
    mod["savings_pct"] = mod["labor_savings"] / mod["total_annual_labor_cost_usd"]
    mod["roi_rank"]    = mod["net_annual_benefit"].rank(ascending=False, method="min").astype(int)
    mod = mod.sort_values("roi_rank")

    for i, (_, row) in enumerate(mod.iterrows()):
        pb_display = f"{row['payback_months']:.0f}" if row["payback_months"] < 1000 else "N/A"
        vals = [
            int(row["roi_rank"]), row["role_name"], row["function"],
            round(row["impl_cost_adj"] / 1e3, 1),
            round(row["net_annual_benefit"] / 1e3, 1),
            f"{row['roi_year1']*100:.0f}%",
            pb_display,
            round(row["npv_3yr"] / 1e3, 1),
            round(row["labor_savings"] / 1e3, 1),
            f"{row['savings_pct']*100:.1f}%",
        ]
        write_data_row(ws, 4 + i, vals, alt=i % 2 == 0)

        # Colour payback cell
        col_pb = 7
        pb_cell = ws.cell(row=4+i, column=col_pb)
        if row["payback_months"] < 12:
            pb_cell.fill = GREEN_FILL
        elif row["payback_months"] < 24:
            pb_cell.fill = AMBER_FILL
        elif row["payback_months"] < 1000:
            pb_cell.fill = RED_FILL

    insert_image(ws, os.path.join(config.CHARTS_DIR, "03_savings_vs_impl_cost.png"), "L2", 540, 300)
    insert_image(ws, os.path.join(config.CHARTS_DIR, "08_roi_vs_complexity.png"), "L22", 540, 300)

    set_col_widths(ws, [6, 30, 18, 16, 20, 16, 14, 16, 18, 18])


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — AI PRIORITY MATRIX
# ══════════════════════════════════════════════════════════════════════════════

def sheet_priority_matrix(wb, scoring, master):
    ws = wb.create_sheet("7. AI Priority Matrix")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    title_block(ws, 1, 1, "AI Priority Matrix — Role Classification & Phase Plan",
                "Quadrant: Auto Potential (x) vs Risk Score (y) | Phase driven by AAPS x risk-adjusted opportunity")

    headers = [
        "Rank", "Role", "Department", "Phase", "RAG",
        "AAPS", "Auto Potential %", "Risk Score",
        "Auto Pool ($M)", "AI Classification", "Quadrant",
    ]
    write_header_row(ws, 3, headers)

    df = scoring.merge(master[["role_id", "automatable_labor_cost_usd", "risk_score",
                               "automation_potential_pct", "ai_classification"]],
                       on="role_id", how="left", suffixes=("", "_m"))
    if "risk_score_m" in df.columns:
        df["risk_score"] = df["risk_score_m"].fillna(df["risk_score"])
    if "ai_classification_m" in df.columns:
        df["ai_classification"] = df["ai_classification_m"].fillna(df.get("ai_classification", ""))

    def quadrant(auto, risk):
        if auto >= 0.50 and risk < 0.40:
            return "[G] Quick Win"
        elif auto >= 0.50:
            return "[A] Careful Deploy"
        elif risk < 0.40:
            return "[B] Augment First"
        else:
            return "[R] Defer / Manual"

    df["Quadrant"] = df.apply(lambda r: quadrant(
        r.get("automation_potential_pct", 0), r.get("risk_score", 0.5)), axis=1)
    df = df.sort_values("AAPS_rank")

    for i, (_, row) in enumerate(df.iterrows()):
        auto_pool = row.get("automatable_labor_cost_usd", 0)
        auto_pct  = row.get("automation_potential_pct", 0)
        risk      = row.get("risk_score", 0)
        cls       = row.get("ai_classification", "")
        vals = [
            int(row["AAPS_rank"]), row["role_name"], row["function"],
            str(row.get("Phase", "")), str(row.get("RAG", "")),
            round(float(row["AAPS"]), 4),
            f"{auto_pct*100:.1f}%",
            round(risk, 3),
            round(float(auto_pool) / 1e6, 2),
            cls,
            row["Quadrant"],
        ]
        write_data_row(ws, 4 + i, vals, alt=i % 2 == 0)
        ws.cell(row=4+i, column=5).fill = rag_fill(str(row.get("RAG", "")))
        ws.cell(row=4+i, column=5).font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

    insert_image(ws, os.path.join(config.CHARTS_DIR, "09_risk_vs_opportunity.png"), "M2", 540, 310)
    insert_image(ws, os.path.join(config.CHARTS_DIR, "06_sensitivity_tornado.png"), "M22", 540, 290)

    set_col_widths(ws, [6, 30, 18, 10, 8, 10, 14, 12, 14, 20, 20])


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 8 — RECOMMENDATIONS & ACTION PLAN
# ══════════════════════════════════════════════════════════════════════════════

def sheet_recommendations(wb, master, scoring, scenarios):
    ws = wb.create_sheet("8. Recommendations")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    title_block(ws, 1, 1, "AI Adoption Recommendations & Action Plan",
                "Each recommendation is directly supported by analytical findings. Approach: Moderate scenario, 24-month phased rollout.")

    headers = [
        "Phase", "Role / Domain", "Department", "Problem / Opportunity",
        "AI Approach", "Auto vs Aug", "Est. Annual Savings",
        "Productivity Gain", "Key Risk", "Complexity", "Timeline",
        "Business Owner",
    ]
    write_header_row(ws, 3, headers)

    recommendations = [
        # -- PHASE 1 (months 1–8) --
        ("Phase 1", "Customer Support Agent (Tier 1)", "Customer Success",
         "320 FTEs handle ~72% repetitive queries; high error rate, high volume",
         "Deploy conversational AI chatbot for FAQ, ticket categorisation, order updates",
         "Full Automation (tier-1 queries) + Augment (complex)",
         "$6.5M–$9M", "35% productivity uplift", "Customer satisfaction risk if AI errors",
         "Medium", "Months 1–6", "VP Customer Experience"),

        ("Phase 1", "AP/AR & Payroll Operations", "Finance",
         "88% task repetitiveness; OCR + AI invoice processing tools are mature",
         "Implement AI-powered invoice processing, payment matching, payroll automation (ADP/Workday AI)",
         "Full Automation",
         "$3M–$4.5M", "30% productivity uplift", "Regulatory accuracy; SOX compliance",
         "Low", "Months 3–8", "CFO / Finance Ops Lead"),

        ("Phase 1", "HR Operations / Admin", "Human Resources",
         "78% repetitive; HRIS automation and HR chatbots are market-ready",
         "Automate data entry, benefits enrolment, tier-1 HR queries via chatbot",
         "Full Automation (admin) + Augment (advisory)",
         "$1.5M–$2M", "25% productivity uplift", "Employee trust if self-service errors",
         "Low", "Months 2–7", "CHRO / HR Ops Director"),

        # -- PHASE 2 (months 9–18) --
        ("Phase 2", "QA / Test Engineer", "Engineering",
         "62% repetitive; AI test generation and regression tools (Copilot, Codium) highly mature",
         "AI-generated test cases, automated regression suites, defect auto-triage",
         "Full Automation (regression) + Augment (exploratory testing)",
         "$4M–$6M", "40% productivity uplift", "Test coverage gaps if over-relying on AI",
         "Medium", "Months 9–14", "VP Engineering / QA Lead"),

        ("Phase 2", "BI / Reporting Analyst", "Data & Analytics",
         "70% repetitive; scheduled reporting fully automatable with AI-BI tools",
         "Automate scheduled reports; AI-generated insights (ThoughtSpot, Power BI Copilot)",
         "Full Automation (scheduled) + Augment (ad-hoc)",
         "$2M–$3M", "30% productivity uplift", "Insight quality if AI misinterprets data",
         "Low", "Months 8–14", "Chief Data Officer"),

        ("Phase 2", "Sales Development Rep (SDR)", "Sales",
         "65% repetitive; AI prospecting, outreach personalisation tools proven at scale",
         "AI lead scoring, automated outreach sequences, CRM auto-logging (Outreach.io, Apollo.io)",
         "Augmentation + Partial Automation",
         "$4M–$6M", "35% productivity uplift", "Brand risk if AI outreach feels impersonal",
         "Medium", "Months 10–18", "VP Sales / Revenue Operations"),

        ("Phase 2", "Content Writer / Copywriter", "Marketing",
         "High AI readiness (0.88); GenAI proven for social, email, blog drafting",
         "AI-assisted content generation (Claude/GPT); human editors focus on strategy",
         "Augmentation",
         "$1.5M–$2.5M", "40% productivity uplift", "Brand voice inconsistency",
         "Low", "Months 8–14", "CMO / Content Lead"),

        # -- PHASE 3 (months 19–24) --
        ("Phase 3", "Software Engineer", "Engineering",
         "High cost pool ($71M); coding assistance (Copilot) already proven — expand org-wide",
         "Mandatory Copilot adoption for all engineers; AI code review, doc generation",
         "Augmentation",
         "$8M–$12M", "35% productivity uplift", "Over-reliance on AI in critical systems",
         "High", "Months 18–24", "CTO / Engineering Leadership"),

        ("Phase 3", "Data Analyst", "Data & Analytics",
         "AI accelerates EDA, cleaning, dashboard creation — frees analysts for strategic work",
         "AI-assisted data cleaning (OpenRefine/AI), NL-to-SQL tools, automated EDA",
         "Augmentation",
         "$3M–$4.5M", "40% productivity uplift", "Data governance and AI bias in insights",
         "Medium", "Months 20–24", "Chief Data Officer"),

        ("Phase 3", "Contract / Paralegal Specialist", "Legal",
         "Contract AI tools (Harvey, Kira) are production-ready; 58% repetitive tasks",
         "AI contract review, legal research acceleration, document management automation",
         "Augmentation (high regulatory oversight required)",
         "$1M–$1.5M", "30% productivity uplift", "Regulatory liability if AI errors missed",
         "High", "Months 20–24", "General Counsel"),
    ]

    phase_fills = {
        "Phase 1": PatternFill("solid", fgColor="D5F5E3"),
        "Phase 2": PatternFill("solid", fgColor="FEF9E7"),
        "Phase 3": PatternFill("solid", fgColor="D6EAF8"),
    }

    for i, rec in enumerate(recommendations):
        write_data_row(ws, 4 + i, list(rec), alt=False)
        phase_cell = ws.cell(row=4+i, column=1)
        phase_cell.fill = phase_fills.get(rec[0], WHITE_FILL)
        phase_cell.font = LABEL_FONT
        ws.row_dimensions[4 + i].height = 36

    # Roles NOT recommended for automation
    title_block(ws, 16, 1, "Roles NOT Recommended for Automation",
                "These roles require human judgment, relationship management, or carry unacceptable AI risk.")
    no_auto_headers = ["Role", "Department", "Reason", "Recommended Approach"]
    write_header_row(ws, 17, no_auto_headers)
    no_auto = [
        ("Corporate Counsel",         "Legal",       "0.92 decision intensity; 0.95 regulatory sensitivity; liability is existential", "Minimal AI (research assist only); human-led"),
        ("Account Executive (Enterprise)", "Sales",  "Customer relationships are the product; AI cannot replicate trust-building",      "CRM automation only; human-led selling"),
        ("HR Business Partner",       "Human Resources", "Employee relations and conflict resolution require empathy and judgment",    "Analytics support; coaching data only"),
        ("FP&A Manager",              "Finance",     "Strategic financial planning requires business context and executive trust",     "AI scenario modelling support; human decides"),
        ("Product Manager",           "Product",     "Product strategy and stakeholder alignment are uniquely human skills",         "AI research synthesis; human roadmap ownership"),
    ]
    for i, row in enumerate(no_auto):
        write_data_row(ws, 18 + i, list(row), alt=i % 2 == 0)
        ws.cell(row=18+i, column=3).fill = PatternFill("solid", fgColor="FADBD8")

    # Assumptions note
    disc_row = 25
    ws.merge_cells(f"A{disc_row}:L{disc_row}")
    c = ws.cell(row=disc_row, column=1,
                value="DISCLAIMER: All data is synthetic. These recommendations are illustrative and must be validated with actual HR, financial, and operational data before any workforce decisions are made.")
    c.font = Font(name="Calibri", size=9, italic=True, color="C0392B")
    c.alignment = LEFT_ALIGN

    set_col_widths(ws, [10, 28, 18, 38, 42, 24, 18, 18, 36, 14, 14, 22])
    for r in range(4, 25):
        ws.row_dimensions[r].height = 40


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    utils.section("08 — EXCEL EXPORT")

    master   = pd.read_csv(os.path.join(config.DATA_PROC, "roles_enriched.csv"))
    scoring  = pd.read_csv(os.path.join(config.DATA_PROC, "scoring_output.csv"))
    scenarios = pd.read_csv(os.path.join(config.DATA_PROC, "scenario_results.csv"))
    summary   = pd.read_csv(os.path.join(config.DATA_PROC, "scenario_summary.csv"))

    # Merge AAPS / Phase / RAG into master
    for col in ["AAPS", "Phase", "RAG", "AAPS_rank", "ai_classification",
                "F1_norm", "F2_norm", "F3_norm", "F4_norm"]:
        if col in scoring.columns and col not in master.columns:
            master = master.merge(scoring[["role_id", col]], on="role_id", how="left")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    print("  Building Sheet 1: Executive Summary...")
    sheet_executive_summary(wb, master, scoring, scenarios, summary)
    print("  Building Sheet 2: Workforce & Cost Analysis...")
    sheet_workforce_cost(wb, master)
    print("  Building Sheet 3: AI Automation Analysis...")
    sheet_ai_automation(wb, master)
    print("  Building Sheet 4: Role & Task Scoring...")
    sheet_role_scoring(wb, scoring)
    print("  Building Sheet 5: Scenario Analysis...")
    sheet_scenario_analysis(wb, scenarios, summary, master)
    print("  Building Sheet 6: ROI & Savings...")
    sheet_roi_savings(wb, scenarios, master)
    print("  Building Sheet 7: AI Priority Matrix...")
    sheet_priority_matrix(wb, scoring, master)
    print("  Building Sheet 8: Recommendations...")
    sheet_recommendations(wb, master, scoring, scenarios)

    wb.save(config.EXCEL_PATH)
    print(f"\n  Excel workbook saved: {config.EXCEL_PATH}")


if __name__ == "__main__":
    main()
