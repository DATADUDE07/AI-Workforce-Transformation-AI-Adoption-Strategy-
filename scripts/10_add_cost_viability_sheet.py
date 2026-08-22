"""
add_cost_viability_sheet.py
----------------------------
Adds the 'AI Cost Viability' sheet to the existing AI_Workforce_Strategy.xlsx.
Run this AFTER 09_ai_cost_viability_analysis.py has been executed.
"""

import os
import sys
import numpy as np
import pandas as pd

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

sys.path.insert(0, r"C:\Users\sahil\.gemini\antigravity\scratch\ai_workforce_analytics")
import config

# ── Style constants (matching 08_excel_export.py) ─────────────────────────────
HEADER_FONT  = Font(name="Calibri", bold=True,  color="FFFFFF",  size=11)
TITLE_FONT   = Font(name="Calibri", bold=True,  size=13, color="1A3C5E")
LABEL_FONT   = Font(name="Calibri", bold=True,  size=10, color="1A3C5E")
BODY_FONT    = Font(name="Calibri", size=10)
SMALL_FONT   = Font(name="Calibri", size=9,  italic=True, color="7F8C8D")
BOLD_FONT    = Font(name="Calibri", bold=True, size=10)

NAVY_FILL    = PatternFill("solid", fgColor="1A3C5E")
GREEN_FILL   = PatternFill("solid", fgColor="27AE60")
AMBER_FILL   = PatternFill("solid", fgColor="F39C12")
RED_FILL     = PatternFill("solid", fgColor="C0392B")
LIGHT_GREEN  = PatternFill("solid", fgColor="D5E8D4")
LIGHT_RED    = PatternFill("solid", fgColor="FADBD8")
LIGHT_AMBER  = PatternFill("solid", fgColor="FDEBD0")
LIGHT_BLUE   = PatternFill("solid", fgColor="D6EAF8")
LIGHT_GREY   = PatternFill("solid", fgColor="F2F3F4")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT_ALIGN  = Alignment(horizontal="right",  vertical="center")

THIN_BORDER  = Border(
    left=Side(style="thin", color="BDC3C7"), right=Side(style="thin", color="BDC3C7"),
    top=Side(style="thin",  color="BDC3C7"), bottom=Side(style="thin", color="BDC3C7"),
)


def _cell(ws, row, col, value, font=None, fill=None, align=None, fmt=None, border=True):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font      = font
    if fill:   c.fill      = fill
    if align:  c.alignment = align
    if fmt:    c.number_format = fmt
    if border: c.border    = THIN_BORDER
    return c


def _merge_title(ws, text, row, col1, col2, fill=NAVY_FILL):
    ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    c = ws.cell(row=row, column=col1, value=text)
    c.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    c.fill      = fill
    c.alignment = CENTER_ALIGN
    return c


def _kpi_box(ws, row, col, value, label, fill_val=NAVY_FILL, fill_lbl=LIGHT_BLUE):
    """Write a 2-row KPI block: big value on top, label below."""
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    v = ws.cell(row=row, column=col, value=value)
    v.font      = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    v.fill      = fill_val
    v.alignment = CENTER_ALIGN
    v.border    = THIN_BORDER

    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
    l = ws.cell(row=row+1, column=col, value=label)
    l.font      = Font(name="Calibri", size=8, color="1A3C5E")
    l.fill      = fill_lbl
    l.alignment = CENTER_ALIGN
    l.border    = THIN_BORDER


def build_cost_viability_sheet(wb, cost_df, sens_df, charts_dir):
    ws = wb.create_sheet("AI Cost Viability")
    ws.sheet_properties.tabColor = "E74C3C"
    ws.freeze_panes = "A8"

    # ── Section Banner ─────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:N1")
    banner = ws["A1"]
    banner.value     = "AI COST VIABILITY ANALYSIS — Is AI Actually Cheaper Than Humans?"
    banner.font      = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    banner.fill      = PatternFill("solid", fgColor="C0392B")
    banner.alignment = CENTER_ALIGN

    ws.merge_cells("A2:N2")
    sub = ws["A2"]
    sub.value     = ("Methodology: Total AI Annual Cost = Ongoing Tooling Cost + (Implementation Cost ÷ 3yr) "
                     "| Human Cost Replaced = Labor Cost × Automation Potential × 65% (Moderate) "
                     "| Cost Efficiency Ratio > 1.0x = AI saves more than it costs")
    sub.font      = SMALL_FONT
    sub.fill      = PatternFill("solid", fgColor="FADBD8")
    sub.alignment = LEFT_ALIGN

    # ── KPI Strip (Row 3-4) ────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 30
    ws.row_dimensions[4].height = 18

    n_cheaper = int(cost_df["ai_is_cheaper"].sum())
    n_total   = len(cost_df)
    cheaper_pct = n_cheaper / n_total * 100

    kpis = [
        (f"{n_cheaper}/{n_total} Roles",   "AI Cheaper Than Human",          NAVY_FILL,   LIGHT_BLUE),
        (f"{n_total-n_cheaper}/{n_total} Roles", "AI Costs More Than It Saves", RED_FILL, LIGHT_RED),
        (f"{cost_df['cost_efficiency_ratio'].min():.2f}x", "Min Cost Efficiency\n(worst role)", AMBER_FILL, LIGHT_AMBER),
        (f"{cost_df['cost_efficiency_ratio'].max():.2f}x", "Max Cost Efficiency\n(best role)",  GREEN_FILL,  LIGHT_GREEN),
        (f"{cost_df['cost_parity_adoption_rate'].min()*100:.1f}%", "Lowest Break-Even\nAdoption Rate", GREEN_FILL, LIGHT_GREEN),
        (f"{cost_df['cost_parity_adoption_rate'].max()*100:.1f}%", "Highest Break-Even\nAdoption Rate", AMBER_FILL, LIGHT_AMBER),
        (f"${cost_df['cost_per_productive_hour'].mean():.0f}/hr", "Avg Human Cost/Hr", NAVY_FILL, LIGHT_BLUE),
        (f"${cost_df['ai_cost_per_hour'].mean():.0f}/hr", "Avg AI Cost/Hr", GREEN_FILL, LIGHT_GREEN),
    ]
    kpi_col = 1
    for val, lbl, fv, fl in kpis:
        _kpi_box(ws, row=3, col=kpi_col, value=val, label=lbl, fill_val=fv, fill_lbl=fl)
        ws.column_dimensions[get_column_letter(kpi_col)].width = 10
        ws.column_dimensions[get_column_letter(kpi_col+1)].width = 5
        kpi_col += 2

    # Blank gap row
    ws.row_dimensions[5].height = 8
    ws.row_dimensions[6].height = 8

    # ── Main Data Table ────────────────────────────────────────────────────────
    headers = [
        "Rank", "Role", "Department", "AI Classification",
        "Human Labor\nReplaced ($)", "Total AI\nAnnual Cost ($)",
        "Net Cost\nPosition ($)", "AI\nCheaper?",
        "Cost Efficiency\nRatio (x)", "Break-Even\nAdoption %",
        "Human\n$/Hr", "AI\n$/Hr", "Ratio\nH vs AI",
        "Verdict"
    ]
    col_widths = [6, 28, 18, 18, 16, 16, 16, 10, 12, 12, 10, 10, 10, 20]
    for i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Section header for table
    _merge_title(ws, "COST VIABILITY DETAIL — All 29 Roles (Sorted by Cost Efficiency, Best to Worst)",
                 row=7, col1=1, col2=len(headers))
    ws.row_dimensions[7].height = 22

    write_row = 8
    write_header_row_local(ws, write_row, headers)
    ws.row_dimensions[write_row].height = 32

    df_sorted = cost_df.sort_values("cost_efficiency_ratio", ascending=False).reset_index(drop=True)

    for rank, (_, row) in enumerate(df_sorted.iterrows(), start=1):
        write_row += 1
        ws.row_dimensions[write_row].height = 18
        alt = (rank % 2 == 0)

        ratio    = round(float(row["cost_efficiency_ratio"]), 2)
        cheaper  = bool(row["ai_is_cheaper"])
        parity   = round(float(row["cost_parity_adoption_rate"]) * 100, 1)
        net_pos  = round(float(row["net_cost_position_usd"]), 0)

        # Verdict text
        if ratio >= 5.0:
            verdict = "OK Excellent — Deploy Immediately"
            verdict_fill = LIGHT_GREEN
        elif ratio >= 2.0:
            verdict = "OK Good — Viable Deployment"
            verdict_fill = LIGHT_GREEN
        elif ratio >= 1.5:
            verdict = "⚠️ Moderate — Monitor Costs"
            verdict_fill = LIGHT_AMBER
        elif ratio >= 1.0:
            verdict = "⚠️ Borderline — High Risk if AI Costs Rise"
            verdict_fill = LIGHT_AMBER
        else:
            verdict = "❌ Unviable — AI More Expensive"
            verdict_fill = LIGHT_RED

        values = [
            rank,
            row["role_name"],
            row["function"],
            row.get("ai_classification", ""),
            round(float(row["human_labor_replaced_usd"]), 0),
            round(float(row["total_ai_annual_cost_usd"]), 0),
            net_pos,
            "YES Y" if cheaper else "NO N",
            ratio,
            parity,
            round(float(row["cost_per_productive_hour"]), 2),
            round(float(row["ai_cost_per_hour"]), 2),
            round(float(row["human_vs_ai_cost_ratio"]), 2),
            verdict,
        ]

        row_fill = LIGHT_GREY if alt else WHITE_FILL
        for col_i, val in enumerate(values, start=1):
            fmt = None
            align = RIGHT_ALIGN if isinstance(val, (int, float)) else LEFT_ALIGN
            if col_i in (5, 6, 7):   fmt = "#,##0"
            if col_i in (9, 13):     fmt = "0.00"
            if col_i == 10:          fmt = '0.0"%"'
            if col_i in (11, 12):    fmt = '"$"#,##0.00'

            c = _cell(ws, write_row, col_i, val,
                      font=BODY_FONT, fill=row_fill, align=align, fmt=fmt)

            # Color the AI Cheaper? column
            if col_i == 8:
                c.fill = LIGHT_GREEN if cheaper else LIGHT_RED
                c.font = Font(name="Calibri", bold=True, size=10,
                              color="1E8449" if cheaper else "922B21")
                c.alignment = CENTER_ALIGN

            # Color the Efficiency Ratio column
            if col_i == 9:
                if ratio >= 2.0:   c.fill = LIGHT_GREEN
                elif ratio >= 1.0: c.fill = LIGHT_AMBER
                else:              c.fill = LIGHT_RED
                c.font = Font(name="Calibri", bold=True, size=10)
                c.alignment = CENTER_ALIGN

            # Color the Net Position column
            if col_i == 7:
                c.fill = LIGHT_GREEN if net_pos >= 0 else LIGHT_RED

            # Color Verdict
            if col_i == 14:
                c.fill = verdict_fill
                c.font = Font(name="Calibri", bold=False, size=9)

    # ── Sensitivity Table ──────────────────────────────────────────────────────
    sens_start_row = write_row + 3
    _merge_title(ws, "AI COST SENSITIVITY — Roles Remaining Viable as AI Ongoing Costs Increase",
                 row=sens_start_row, col1=1, col2=6)

    sens_header = ["AI Cost Increase", "Roles: AI Saves Money", "Roles: AI Too Expensive",
                   "% Portfolio Viable", "Risk Level", "Management Action"]
    write_header_row_local(ws, sens_start_row + 1, sens_header, start_col=1)

    risk_map = {
        "+0%":   ("[G] No Risk",      LIGHT_GREEN,  "No action needed"),
        "+10%":  ("[G] No Risk",      LIGHT_GREEN,  "No action needed"),
        "+25%":  ("[G] Low Risk",     LIGHT_GREEN,  "Monitor quarterly"),
        "+50%":  ("[A] Low-Medium",   LIGHT_AMBER,  "Review contracts with AI vendors"),
        "+75%":  ("[A] Medium Risk",  LIGHT_AMBER,  "Renegotiate or source alternative tools"),
        "+100%": ("[A] Medium Risk",  LIGHT_AMBER,  "Audit highest-cost AI deployments"),
        "+150%": ("[R] High Risk",    LIGHT_RED,    "Pause deployment; full cost review"),
        "+200%": ("[R] High Risk",    LIGHT_RED,    "Re-evaluate business case; consider insourcing"),
    }

    for i, (_, srow) in enumerate(sens_df.iterrows()):
        r_num = sens_start_row + 2 + i
        inc   = str(srow.get("AI Cost Increase (%)", srow.iloc[0]))
        viable = int(srow.get("Roles Where AI Saves Money", srow.iloc[1]))
        too_exp = int(srow.get("Roles Where AI Too Expensive", srow.iloc[2]))
        pct_viable = viable / n_total * 100
        risk_txt, risk_fill, action = risk_map.get(inc, ("—", WHITE_FILL, "—"))

        row_vals = [inc, viable, too_exp, f"{pct_viable:.0f}%", risk_txt, action]
        for ci, val in enumerate(row_vals, start=1):
            c = _cell(ws, r_num, ci, val, font=BODY_FONT, align=CENTER_ALIGN if ci != 6 else LEFT_ALIGN)
            if ci == 5: c.fill = risk_fill

    # ── Embed Charts ───────────────────────────────────────────────────────────
    chart_configs = [
        ("10_ai_cost_efficiency_ratio.png",  f"A{sens_start_row + 14}", 650, 380),
        ("11_cost_parity_threshold.png",      f"H{sens_start_row + 14}", 650, 380),
        ("12_ai_cost_sensitivity.png",        f"A{sens_start_row + 40}", 500, 300),
        ("13_human_vs_ai_cost_per_hour.png",  f"H{sens_start_row + 40}", 550, 330),
    ]
    for fname, anchor, w, h in chart_configs:
        fpath = os.path.join(charts_dir, fname)
        if os.path.exists(fpath):
            img = XLImage(fpath)
            img.width  = w
            img.height = h
            ws.add_image(img, anchor)

    # ── Insight Box ────────────────────────────────────────────────────────────
    insight_row = sens_start_row + 11
    ws.merge_cells(start_row=insight_row, start_column=1,
                   end_row=insight_row + 1, end_column=6)
    ins = ws.cell(row=insight_row, column=1,
                  value=("[!] KEY INSIGHT: Even a +200% increase in AI ongoing costs (e.g., LLM API price hike, license inflation) "
                         "only makes 1 role unviable (Corporate Counsel — already recommended DEFER). "
                         "The portfolio is highly resilient to AI cost inflation because most roles have Efficiency Ratios of 3x–7x."))
    ins.font      = Font(name="Calibri", bold=True, size=10, color="1A3C5E")
    ins.fill      = LIGHT_BLUE
    ins.alignment = LEFT_ALIGN
    ins.border    = THIN_BORDER

    print("  OK 'AI Cost Viability' sheet built successfully.")
    return ws


def write_header_row_local(ws, row_num, headers, start_col=1):
    for j, h in enumerate(headers, start=start_col):
        c = ws.cell(row=row_num, column=j, value=h)
        c.font      = HEADER_FONT
        c.fill      = NAVY_FILL
        c.alignment = CENTER_ALIGN
        c.border    = THIN_BORDER


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    DATA_PROC  = os.path.join(r"C:\Users\sahil\.gemini\antigravity\scratch\ai_workforce_analytics", "dataset", "processed")
    CHARTS_DIR = os.path.join(r"C:\Users\sahil\.gemini\antigravity\scratch\ai_workforce_analytics", "outputs", "charts")
    EXCEL_PATH = config.EXCEL_PATH

    print(f"\nLoading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    # Remove sheet if it already exists (for re-runs)
    if "AI Cost Viability" in wb.sheetnames:
        del wb["AI Cost Viability"]
        print("  Removed existing 'AI Cost Viability' sheet.")

    print("  Loading cost viability data...")
    cost_df = pd.read_csv(os.path.join(DATA_PROC, "ai_cost_viability.csv"))
    sens_df = pd.read_csv(os.path.join(DATA_PROC, "ai_cost_sensitivity.csv"))

    # Add ai_classification from roles_enriched if not present
    if "ai_classification" not in cost_df.columns:
        master = pd.read_csv(os.path.join(DATA_PROC, "roles_enriched.csv"))
        cost_df = cost_df.merge(master[["role_id","ai_classification"]], on="role_id", how="left")

    print("  Building 'AI Cost Viability' sheet...")
    build_cost_viability_sheet(wb, cost_df, sens_df, CHARTS_DIR)

    wb.save(EXCEL_PATH)
    print(f"\n  OK Excel updated and saved: {EXCEL_PATH}")
    print(f"  Sheets: {wb.sheetnames}")
